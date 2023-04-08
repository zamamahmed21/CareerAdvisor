import requests
from bs4 import BeautifulSoup
import openpyxl
from time import sleep

workbook = openpyxl.load_workbook('personality careers.xlsx')

def scrape_data_for_one_p_type(url, sheet_name, next_cursor = ''):
    """scrape data for one personality type from the link provided and save it into an xlsx file.
    Note: If you are passing next_cursor value, include next_cursor value within the link & 
    provide it as a parameter as well.
    """

    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name)

    while True:
        print(url)
        response = requests.get(url)
        if response.status_code!=200:
            print("Server returned the following status error: "+str(response.status_code))
            break
        contents = response.json()

        url = url.replace(f'&nextCursor={next_cursor}','')
        next_cursor = contents['data']['cursor']['nextCursor']
        ppl_data = contents['data']['results']

        for person in ppl_data:
            worksheet.append([person['name'],person['subcategory']])

        if next_cursor=='':
            break

        url = url + f'&nextCursor={next_cursor}'
    
    workbook.save('personality careers.xlsx')


apis_ptypes = [['https://api.personality-database.com/api/v2/types/1/profiles?destination=istj-famous-people&limit=100','ISTJ'],
                  ['https://api.personality-database.com/api/v2/types/2/profiles?destination=estj-famous-people&limit=100','ESTJ'],
                  ['https://api.personality-database.com/api/v2/types/3/profiles?destination=isfj-famous-people&limit=100','ISFJ'],
                  ['https://api.personality-database.com/api/v2/types/4/profiles?destination=esfj-famous-people&limit=100','ESFJ'],
                  ['https://api.personality-database.com/api/v2/types/5/profiles?destination=esfp-famous-people&limit=100','ESFP'],
                  ['https://api.personality-database.com/api/v2/types/6/profiles?destination=isfp-famous-people&limit=100','ISFP'],
                  ['https://api.personality-database.com/api/v2/types/7/profiles?destination=estp-famous-people&limit=100','ESTP'],
                  ['https://api.personality-database.com/api/v2/types/8/profiles?destination=istp-famous-people&limit=100','ISTP'],
                  ['https://api.personality-database.com/api/v2/types/9/profiles?destination=infj-famous-people&limit=100','INFJ'],
                  ['https://api.personality-database.com/api/v2/types/10/profiles?destination=enfj-famous-people&limit=100','ENFJ'],
                  ['https://api.personality-database.com/api/v2/types/11/profiles?destination=infp-famous-people&limit=100','INFP'],
                  ['https://api.personality-database.com/api/v2/types/12/profiles?destination=enfp-famous-people&limit=100','ENFP'],
                  ['https://api.personality-database.com/api/v2/types/13/profiles?destination=intp-famous-people&limit=100','INTP'],
                  ['https://api.personality-database.com/api/v2/types/14/profiles?destination=entp-famous-people&limit=100','ENTP'],
                  ['https://api.personality-database.com/api/v2/types/15/profiles?destination=intj-famous-people&limit=100','INTJ'],
                  ['https://api.personality-database.com/api/v2/types/16/profiles?destination=entj-famous-people&limit=100','ENTJ']
                  ]

for api_ptype in apis_ptypes:
    scrape_data_for_one_p_type(api_ptype[0],api_ptype[1])

# scrape_data_for_one_p_type('https://api.personality-database.com/api/v2/types/7/profiles?destination=estp-famous-people&limit=100&nextCursor=MTUwMA==','ESTP',next_cursor='MTUwMA==')